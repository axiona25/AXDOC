"""Export protocolli Excel/PDF (FASE 25)."""
import io
from datetime import timedelta

import pytest
from django.contrib.auth import get_user_model
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from apps.organizations.models import OrganizationalUnit, OrganizationalUnitMembership
from apps.protocols.models import Protocol

User = get_user_model()


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def ou_it(db):
    return OrganizationalUnit.objects.create(name="IT", code="IT", description="")


@pytest.fixture
def user_admin(db):
    u = User.objects.create_user(email="admin_ex@test.com", password="test")
    u.role = "ADMIN"
    u.save()
    return u


@pytest.fixture
def user_guest(db):
    u = User.objects.create_user(email="guest_ex@test.com", password="test")
    u.user_type = "guest"
    u.save()
    return u


@pytest.mark.django_db
class TestProtocolExport:
    def test_export_excel_returns_200_xlsx(self, api_client, user_admin, ou_it):
        Protocol.objects.create(
            organizational_unit=ou_it,
            year=2025,
            number=1,
            protocol_id="2025/IT/0001",
            direction="in",
            subject="Test export",
            sender_receiver="ACME",
            status="active",
        )
        api_client.force_authenticate(user=user_admin)
        r = api_client.get("/api/protocols/export_excel/")
        assert r.status_code == status.HTTP_200_OK
        assert "spreadsheetml" in r["Content-Type"]
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.active.cell(row=4, column=1).value == "N. Protocollo"

    def test_export_pdf_returns_200_pdf(self, api_client, user_admin, ou_it):
        Protocol.objects.create(
            organizational_unit=ou_it,
            year=2025,
            number=2,
            protocol_id="2025/IT/0002",
            direction="out",
            subject="PDF",
            status="active",
        )
        api_client.force_authenticate(user=user_admin)
        r = api_client.get("/api/protocols/export_pdf/")
        assert r.status_code == status.HTTP_200_OK
        assert r["Content-Type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"

    def test_export_filtered_by_date(self, api_client, user_admin, ou_it):
        p_old = Protocol.objects.create(
            organizational_unit=ou_it,
            year=2025,
            number=10,
            protocol_id="2025/IT/0010",
            direction="in",
            subject="Vecchio",
            status="active",
        )
        p_new = Protocol.objects.create(
            organizational_unit=ou_it,
            year=2025,
            number=11,
            protocol_id="2025/IT/0011",
            direction="in",
            subject="Nuovo",
            status="active",
        )
        old = timezone.now() - timedelta(days=60)
        Protocol.objects.filter(pk=p_old.pk).update(created_at=old)
        Protocol.objects.filter(pk=p_new.pk).update(created_at=timezone.now())

        api_client.force_authenticate(user=user_admin)
        cutoff = (timezone.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        r = api_client.get("/api/protocols/export_excel/", {"date_from": cutoff})
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        found_new = False
        for row in range(5, ws.max_row + 1):
            if ws.cell(row=row, column=4).value == "Nuovo":
                found_new = True
        assert found_new
        found_old = False
        for row in range(5, ws.max_row + 1):
            if ws.cell(row=row, column=4).value == "Vecchio":
                found_old = True
        assert not found_old

    def test_guest_cannot_export(self, api_client, user_guest):
        api_client.force_authenticate(user=user_guest)
        r = api_client.get("/api/protocols/export_excel/")
        assert r.status_code == status.HTTP_403_FORBIDDEN
