"""Export documenti Excel (FASE 25)."""
import io

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from apps.documents.models import Document, Folder

User = get_user_model()


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def user(db):
    u = User.objects.create_user(email="doc_ex@test.com", password="test", role="OPERATOR")
    return u


@pytest.mark.django_db
class TestDocumentExport:
    def test_export_documents_excel_200(self, api_client, user):
        folder = Folder.objects.create(name="F", created_by=user)
        Document.objects.create(
            title="DocExport",
            folder=folder,
            status=Document.STATUS_DRAFT,
            created_by=user,
        )
        api_client.force_authenticate(user=user)
        r = api_client.get("/api/documents/export_excel/")
        assert r.status_code == status.HTTP_200_OK
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.active.cell(row=4, column=1).value == "Titolo"

    def test_export_documents_filtered_by_status(self, api_client, user):
        folder = Folder.objects.create(name="F2", created_by=user)
        Document.objects.create(
            title="Bozza",
            folder=folder,
            status=Document.STATUS_DRAFT,
            created_by=user,
        )
        Document.objects.create(
            title="Approvato",
            folder=folder,
            status=Document.STATUS_APPROVED,
            created_by=user,
        )
        api_client.force_authenticate(user=user)
        r = api_client.get("/api/documents/export_excel/", {"status": "APPROVED"})
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        titles = [wb.active.cell(row=r, column=1).value for r in range(5, wb.active.max_row + 1)]
        assert "Approvato" in titles
        assert "Bozza" not in titles
