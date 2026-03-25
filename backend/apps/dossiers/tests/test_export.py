"""Export fascicoli (FASE 25)."""
import io

import pytest
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from django.contrib.auth import get_user_model

from apps.dossiers.models import Dossier

User = get_user_model()


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def user_admin(db):
    u = User.objects.create_user(email="admin_de@test.com", password="test")
    u.role = "ADMIN"
    u.save()
    return u


@pytest.mark.django_db
def test_export_dossiers_excel_200(api_client, user_admin):
    Dossier.objects.create(
        title="F1",
        identifier="F-EXP-1",
        status="open",
        created_by=user_admin,
    )
    api_client.force_authenticate(user=user_admin)
    r = api_client.get("/api/dossiers/export_excel/", {"filter": "all"})
    assert r.status_code == status.HTTP_200_OK
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.active.cell(row=4, column=1).value == "Codice"
