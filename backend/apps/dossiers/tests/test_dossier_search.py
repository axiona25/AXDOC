"""Ricerca testuale lista fascicoli ed export (FASE 37)."""
from io import BytesIO

import pytest
from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient

from apps.dossiers.models import Dossier
from apps.organizations.models import OrganizationalUnit, OrganizationalUnitMembership, Tenant

User = get_user_model()


@pytest.fixture
def tenant(db):
    t, _ = Tenant.objects.get_or_create(
        slug="default",
        defaults={"name": "Organizzazione Default", "plan": "enterprise"},
    )
    return t


@pytest.fixture
def ou(db, tenant):
    return OrganizationalUnit.objects.create(name="DS OU", code="DS", tenant=tenant)


@pytest.fixture
def admin_user(db, tenant, ou):
    u = User.objects.create_user(email="ds-search-adm@test.com", password="Test123!", role="ADMIN")
    u.tenant = tenant
    u.save(update_fields=["tenant"])
    OrganizationalUnitMembership.objects.create(user=u, organizational_unit=ou, role="ADMIN")
    return u


@pytest.fixture
def admin_client(admin_user):
    c = APIClient()
    c.force_authenticate(user=admin_user)
    return c


@pytest.mark.django_db
def test_dossier_search_by_title(admin_client, tenant, ou, admin_user):
    d = Dossier.objects.create(
        title="Fascicolo ricerca titolo DS",
        identifier="DS-T-1",
        tenant=tenant,
        created_by=admin_user,
        organizational_unit=ou,
    )
    r = admin_client.get("/api/dossiers/", {"search": "ricerca titolo"})
    assert r.status_code == 200
    data = r.json()
    rows = data.get("results", data) if isinstance(data, dict) else data
    ids = [x["id"] for x in rows] if isinstance(rows, list) else []
    assert str(d.id) in ids


@pytest.mark.django_db
def test_dossier_search_by_identifier(admin_client, tenant, ou, admin_user):
    d = Dossier.objects.create(
        title="X",
        identifier="UNIQ-ID-DS-42",
        tenant=tenant,
        created_by=admin_user,
        organizational_unit=ou,
    )
    r = admin_client.get("/api/dossiers/", {"search": "UNIQ-ID-DS-42"})
    assert r.status_code == 200
    data = r.json()
    rows = data.get("results", data) if isinstance(data, dict) else data
    assert any(x["id"] == str(d.id) for x in rows)


@pytest.mark.django_db
def test_dossier_search_empty_returns_all(admin_client, tenant, ou, admin_user):
    Dossier.objects.create(
        title="A DS",
        identifier="DS-A-1",
        tenant=tenant,
        created_by=admin_user,
        organizational_unit=ou,
    )
    r = admin_client.get("/api/dossiers/", {})
    assert r.status_code == 200
    c1 = len(r.json().get("results", []))
    r2 = admin_client.get("/api/dossiers/", {"search": ""})
    assert r2.status_code == 200
    assert len(r2.json().get("results", [])) == c1


@pytest.mark.django_db
def test_dossier_export_respects_search_filter(admin_client, tenant, ou, admin_user):
    Dossier.objects.create(
        title="IncludeMeDSExport",
        identifier="E1",
        tenant=tenant,
        created_by=admin_user,
        organizational_unit=ou,
    )
    Dossier.objects.create(
        title="Other",
        identifier="E2",
        tenant=tenant,
        created_by=admin_user,
        organizational_unit=ou,
    )
    r = admin_client.get("/api/dossiers/export_excel/", {"search": "IncludeMeDSExport"})
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    flat = " ".join(str(c or "") for row in wb.active.iter_rows(values_only=True) for c in row)
    assert "IncludeMeDSExport" in flat
    assert "Other" not in flat
