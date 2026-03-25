"""Export audit log (FASE 25)."""
import io
from datetime import timedelta

import pytest
from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from apps.authentication.models import AuditLog

User = get_user_model()


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def admin(db):
    u = User.objects.create_user(email="admin_au@test.com", password="test")
    u.role = "ADMIN"
    u.save()
    return u


@pytest.fixture
def operator(db):
    return User.objects.create_user(email="op_au@test.com", password="test", role="OPERATOR")


@pytest.mark.django_db
class TestAuditExport:
    def test_export_audit_excel_admin_only(self, api_client, admin, operator):
        AuditLog.log(admin, "LOGIN", {}, None)
        api_client.force_authenticate(user=operator)
        r = api_client.get("/api/audit/export_excel/")
        assert r.status_code == status.HTTP_403_FORBIDDEN

        api_client.force_authenticate(user=admin)
        r = api_client.get("/api/audit/export_excel/")
        assert r.status_code == status.HTTP_200_OK
        assert "spreadsheetml" in r["Content-Type"]

    def test_export_audit_filtered_by_date(self, api_client, admin):
        old = timezone.now() - timedelta(days=90)
        AuditLog.objects.create(user=admin, action="LOGIN", detail={})
        new_log = AuditLog.objects.create(user=admin, action="LOGOUT", detail={})
        AuditLog.objects.filter(pk=new_log.pk).update(timestamp=timezone.now())
        AuditLog.objects.exclude(pk=new_log.pk).update(timestamp=old)

        api_client.force_authenticate(user=admin)
        cutoff = (timezone.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        r = api_client.get("/api/audit/export_excel/", {"date_from": cutoff})
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        actions = [ws.cell(row=r, column=3).value for r in range(5, ws.max_row + 1)]
        assert "LOGOUT" in actions
        assert "LOGIN" not in actions
