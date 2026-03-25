"""Test ExportService (FASE 25)."""
import io

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from apps.dashboard.export_service import ExportService


@pytest.mark.django_db
class TestExportService:
    def test_generate_excel_returns_xlsx_response(self):
        resp = ExportService.generate_excel(
            title="Test Report",
            headers=["A", "B"],
            rows=[[1, 2]],
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp["Content-Type"]
        assert resp.content[:2] == b"PK"

    def test_generate_pdf_returns_pdf_response(self):
        resp = ExportService.generate_pdf(
            title="Test PDF",
            headers=["Col1"],
            rows=[["x"]],
            orientation="portrait",
        )
        assert resp.status_code == 200
        assert resp["Content-Type"] == "application/pdf"
        assert resp.content[:4] == b"%PDF"

    def test_excel_has_correct_headers(self):
        resp = ExportService.generate_excel(
            title="H",
            headers=["Nome", "Valore"],
            rows=[],
        )
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.cell(row=4, column=1).value == "Nome"
        assert ws.cell(row=4, column=2).value == "Valore"

    def test_pdf_landscape_orientation(self):
        resp = ExportService.generate_pdf(
            title="Wide",
            headers=["a", "b", "c"],
            rows=[[1, 2, 3]],
            orientation="landscape",
        )
        buf = io.BytesIO(resp.content)
        reader = PdfReader(buf)
        page = reader.pages[0]
        w = float(page.mediabox.width)
        h = float(page.mediabox.height)
        assert w > h
