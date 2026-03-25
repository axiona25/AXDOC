import io
import os
from datetime import datetime

from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class ExportService:
    """Servizio generico per esportazione dati in Excel e PDF."""

    @staticmethod
    def generate_excel(
        title: str,
        headers: list[str],
        rows: list[list],
        sheet_name: str = "Report",
        column_widths: list[int] | None = None,
    ) -> HttpResponse:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title_font = Font(name="Arial", bold=True, size=14)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ncols = max(len(headers), 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = title_font

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=2, column=1, value=f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row_data in enumerate(rows, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        widths = column_widths or [18] * len(headers)
        for idx in range(len(headers)):
            letter = get_column_letter(idx + 1)
            ws.column_dimensions[letter].width = widths[idx] if idx < len(widths) else 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def generate_pdf(
        title: str,
        headers: list[str],
        rows: list[list],
        orientation: str = "portrait",
        column_widths: list[float] | None = None,
        subtitle: str = "",
    ) -> HttpResponse:
        buffer = io.BytesIO()
        pagesize = landscape(A4) if orientation == "landscape" else A4

        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=20 * mm,
            bottomMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=6,
            textColor=colors.HexColor("#2E75B6"),
        )
        subtitle_style = ParagraphStyle(
            "ReportSubtitle",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=12,
        )
        cell_style = ParagraphStyle(
            "CellStyle",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
        )

        elements = []
        logo_path = getattr(settings, "EXPORT_REPORT_LOGO", None)
        if logo_path and os.path.isfile(logo_path):
            try:
                elements.append(Image(logo_path, width=35 * mm, height=12 * mm))
                elements.append(Spacer(1, 3 * mm))
            except OSError:
                pass

        elements.append(Paragraph(title, title_style))
        sub = subtitle or f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        elements.append(Paragraph(sub, subtitle_style))
        elements.append(Spacer(1, 6 * mm))

        table_data = [headers]
        for row in rows:
            table_data.append(
                [Paragraph(str(v) if v is not None else "", cell_style) for v in row]
            )

        if column_widths:
            col_w = [w * mm for w in column_widths]
        else:
            available = pagesize[0] - 30 * mm
            col_w = [available / len(headers)] * len(headers) if headers else [available]

        table = Table(table_data, colWidths=col_w, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 0), (-1, 0), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 1), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
                ]
            )
        )

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
